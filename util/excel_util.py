# _*_ coding: utf-8 _*_

"""
excel处理类
"""


from openpyxl import Workbook
from openpyxl import load_workbook

class ExcelUtil(object):

    def __init__(self,file_name,sheet_name=None,sheet_index=0):
        self.file_name = file_name
        self.sheet_name = sheet_name
        self.sheet_index = sheet_index

    def create_excel(self,data_dict):
        """创建excel"""
        workbook = Workbook()
        sheet = None
        if self.sheet_name is None:
            sheet = workbook.create_sheet(index=self.sheet_index)
        else:
            sheet = workbook.create_sheet(title=self.sheet_name,index=self.sheet_index)
        for key in data_dict:
            sheet[key]=data_dict[key]
        workbook.save(self.file_name)
        workbook.close()

    def load_excel(self):
        """加载excel"""
        self.workbook = load_workbook(self.file_name)
        self.sheet = None
        if self.sheet_name is None:
            self.sheet = self.workbook.worksheets[self.sheet_index]
        else:
            self.sheet = self.workbook[self.sheet_name]

    def get_data_by_cell_no(self,cell_no):
        """根据编号获取单元格的值"""
        self.load_excel()
        return self.sheet[cell_no].value

    def get_data_by_row_col_no(self,row_no,col_no):
        """根据行、列索引获取单元格的值"""
        self.load_excel()
        return self.sheet.cell(row_no,col_no).value

    def set_data_by_cell_no(self,cell_no,cell_value):
        """根据编号修改单元格的值"""
        self.load_excel()
        self.sheet[cell_no]=cell_value
        self.workbook.save(self.file_name)

    def set_data_by_row_col_no(self,row_no,col_no,cell_value):
        """根据行、列索引修改单元格的值"""
        self.load_excel()
        self.sheet.cell(row_no,col_no,cell_value)
        self.workbook.save(self.file_name)

    def get_data_by_row_no(self, row):
        """获取整行的所有值，传入行数"""
        self.load_excel()
        columns = self.sheet.max_column
        rowdata = []
        for i in range(1, columns + 1):
            cellvalue = self.sheet.cell(row=row, column=i).value
            rowdata.append(cellvalue)
        return rowdata

    def get_data_by_col_no(self,col):
        """获取整列的所有值，存入列号"""
        self.load_excel()
        rows = self.sheet.max_row
        coldata = []
        for i in range(1, rows + 1):
            cellvalue = self.sheet.cell(row=i, column=col).value
            coldata.append(cellvalue)
        return coldata

    def get_row_no_by_cell_value(self,cell_value,col_no):
        """根据单元格值获取行号"""
        self.load_excel()
        rowns = self.sheet.max_row
        row_no = -1
        for i in range(rowns):
            if self.sheet.cell(i+1,col_no).value == cell_value:
                row_no = i+1
                break
            else:
                continue
        return row_no

    def get_all_dict_data(self):
        """读取表格数据转换成字典的列表格式显示"""
        keys = self.get_data_by_row_no(1)
        rowNum = self.sheet.max_row  # 获取总行数
        colNum = self.sheet.max_column  # 获取总列数
        if rowNum < 1:
            print("总行数小于 1")
        else:
            result = []
            j = 2
            for i in range(rowNum - 1):
                s = {}
                # 从第二行取对应 values 值
                values = self.get_data_by_row_no(j)
                for x in range(colNum):
                    s[keys[x]] = values[x]
                result.append(s)
                j += 1
        return result


    def get_case_list(self):
        """获取所有用例数据的列表，每个用例数据按{"行号":[用例数据]}存储"""
        case_data_list = []
        row_counts = self.sheet.max_row
        for row_no in range(row_counts-1):
            row_no = row_no + 2
            case_list = self.get_data_by_row_no(row_no)
            case_dict = {row_no:case_list}
            case_data_list.append(case_dict)
        return case_data_list
